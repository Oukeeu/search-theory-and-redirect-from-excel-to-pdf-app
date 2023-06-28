import pandas as pd
import PyPDF2
import openpyxl
import sys, os
from PyQt5.QtWidgets import QApplication, QWidget, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit, QTextEdit, QPushButton
from PyQt5.QtWidgets import QApplication, QFileDialog
from PyQt5.QtGui import QDesktopServices
from PyQt5.QtCore import QUrl

class MainWindow(QWidget):
    def __init__(self):
        super().__init__()

        self.setStyleSheet("background-color: #f5f5f5;")
        self.setFixedSize(500, 400)

        # Создаем виджеты
        self.query_label = QLabel("Введите запрос:")
        self.query_edit = QLineEdit()
        self.response_label = QLabel("Ответ:")
        self.response_edit = QTextEdit()
        self.cross_button = QPushButton("Перейти к файлу")
        self.send_button = QPushButton("Отправить")

        # Создаем вертикальный и горизонтальный layout'ы
        v_layout = QVBoxLayout()
        h_layout = QHBoxLayout()

        # Добавляем виджеты в layout'ы
        h_layout.addWidget(self.query_label)
        h_layout.addWidget(self.query_edit)
        v_layout.addLayout(h_layout)

        v_layout.addWidget(self.response_label)
        v_layout.addWidget(self.response_edit)

        v_layout.addWidget(self.cross_button)
        v_layout.addWidget(self.send_button)

        # Устанавливаем layout для окна
        self.setLayout(v_layout)
        self.send_button.clicked.connect(self.send_query)
        self.cross_button.clicked.connect(self.cross_query)

    def find_page_with_max_match(self, pdf_reader, keywords):
        max_match = 0
        max_match_page_num = -1

        for page in range(len(pdf_reader.pages)):
            page_obj = pdf_reader.pages[page]
            text = page_obj.extract_text()
            match_count = 0
            for keyword in keywords:
                if keyword in text:
                    match_count += 1
            if match_count > max_match:
                max_match = match_count
                max_match_page_num = page

        return max_match_page_num

    def cross_query(self):
        keywords = self.response_edit.toPlainText()

        keywords = keywords.split(" ")

        pdf_file = open('gost_21672-99.pdf', 'rb')
        pdf_reader = PyPDF2.PdfReader(pdf_file)

        max_match_page_num = self.find_page_with_max_match(pdf_reader, keywords)
        if max_match_page_num != -1:

            executable_path = os.path.abspath(__file__)
            file_dir = os.path.dirname(executable_path)

            url = 'file:///' + file_dir + '/gost_21672-99.pdf#page=' + str(max_match_page_num + 1)
            print('Совпадения на странице: ', max_match_page_num + 1)

            QDesktopServices.openUrl(QUrl.fromLocalFile(url))
        else:
            print('Совпадений не найдено')
        pdf_file.close()


    def send_query(self):
        query = self.query_edit.text()

        illuminator_type = query

        # открываем файл Excel
        workbook = openpyxl.load_workbook('Vse_illyuminatory.xlsx')

        # перебираем все листы
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]

            # проходим по строкам и ищем нужную строку
            for row in worksheet.iter_rows():  # проверяем условие для ячейки
                if row[1].value == illuminator_type:  # выводим найденную строку в консоль
                    string = ''
                    for cell in row:
                        if (cell.value != None and cell.value != '-' and cell.value != '+' and cell.value != ''):
                            string += str(cell.value) + ' '

        self.response_edit.setText(f"{string}")

if __name__ == '__main__':
    app = QApplication([])
    window = MainWindow()
    window.show()
    app.exec_()
